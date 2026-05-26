"""
tabla_loader.py
Carga y normaliza la tabla de composición de alimentos.

La tabla tiene estructura fija en todas las hojas:
  Fila 0: encabezados de sección (COMPOSICION, VITAMINAS, etc.)
  Fila 1: sub-encabezados (unidades)
  Fila 2: sub-encabezados (unidades)
  Fila 3: nombres de columnas (NRO, ALIMENTO, ESTADO, CAL, PR, ...)
  Fila 4+: datos

Columnas en posición (índice base 0):
  0: NRO     — número de ítem
  1: ALIMENTO
  2: ESTADO  — estado/preparación
  3: CAL     — Energía (kcal/100g)
  4: PR      — Proteínas (g/100g)
  5: GR      — Grasas totales (g/100g)
  6: HC      — Hidratos de carbono (g/100g)
  7: H2O     — Agua (g/100g)
  8: NE      — Nitrógeno equivalente
  9: VIT_A   — Vitamina A (UI/100g)
 10: VIT_B1  — Vitamina B1/Tiamina (mcg/100g)
 11: VIT_B2  — Vitamina B2/Riboflavina (mcg/100g)
 12: VIT_C   — Vitamina C (mcg/100g)
 13: NIAC    — Niacina (mcg/100g)
 14: NA      — Sodio (mg/100g)
 15: K       — Potasio (mg/100g)
 16: CA      — Calcio (mg/100g)
 17: MG      — Magnesio (mg/100g)
 18: FE      — Hierro (mg/100g)
 19: CU      — Cobre (mg/100g)
 20: P       — Fósforo (mg/100g)
 21: S       — Azufre (mg/100g)
 22: CL      — Cloro (mg/100g)
 23-30: Aminoácidos (Fen, Ileu, Leu, Lis, Met, Tre, Tri, Val) mg/100g
 31: ACID    — Acidez
 32: ALCAL   — Alcalinidad
"""

import pandas as pd
import numpy as np
import openpyxl
import re
from pathlib import Path

# ─── Mapeo de hojas a grupo lógico ────────────────────────────────────────────
HOJAS_GRUPOS = {
    'Cereales':                 'Cereales',
    'leguminosas':              'Leguminosas',
    'tuberculos y hortalizas':  'Hortalizas',
    'frutos frescos':           'Frutas',
    'frutos secos':             'FrutosSecos',
    'Leche y derivados':        'Lacteos',
    'Huevos':                   'Huevos',
    'Azucares y dulces varios': 'Azucares',
    'aceites y grasas':         'Aceites',
    'Pescados':                 'Pescados',
    'Carne':                    'Carnes',
    'Cerdo':                    'Carnes',
    'Cordero':                  'Carnes',
    'Oveja':                    'Carnes',
    'Ternera':                  'Carnes',
    'Vaca':                     'Carnes',
    'Embutidos':                'Embutidos',
    'Aves':                     'Aves',
    'Caza':                     'Carnes',
}

COL_NAMES = [
    'NRO', 'ALIMENTO', 'ESTADO',
    'CAL', 'PR', 'GR', 'HC', 'H2O', 'NE',
    'VIT_A', 'VIT_B1', 'VIT_B2', 'VIT_C', 'NIAC',
    'NA', 'K', 'CA', 'MG', 'FE', 'CU', 'P', 'S', 'CL',
    'FEN', 'ILEU', 'LEU', 'LIS', 'MET', 'TRE', 'TRI', 'VAL',
    'ACID', 'ALCAL'
]

NUTRIENT_COLS = [
    'CAL', 'PR', 'GR', 'HC',
    'VIT_A', 'VIT_B1', 'VIT_B2', 'VIT_C',
    'CA', 'FE', 'MG', 'K', 'P'
]

# ─── FIX 1: Reclasificaciones post-carga ──────────────────────────────────────
# Alimentos mal ubicados en la hoja de origen que deben ir a otro grupo lógico
RECLASIFICACIONES = {
    # (alimento_lower, estado_lower_o_None) → grupo_correcto
    ('soja', 'harina'):   'Leguminosas',
    ('soja', None):       'Leguminosas',    # cualquier estado de soja
}

# ─── FIX 2: Lista de exclusión de disponibilidad ──────────────────────────────
# Alimentos que existen en la tabla nutricional pero NO son de compra cotidiana
# en supermercados/almacenes argentinos, o son ingredientes industriales.
# Se marcan DISPONIBLE=False y el optimizador no los usa.
#
# Criterios de exclusión:
#   A) Fauna no disponible en comercios (caza exótica, ballena)
#   B) Ingredientes industriales (almidones puros, extractos)
#   C) Formas concentradas/deshidratadas de alimentos frescos básicos
#      (huevo seco, leche total seca — la leche en polvo descremada sí es común)
#   D) Productos no identificables en góndola argentina (majuela, achicoria tostada*)
#
# * achicoria tostada es sustituto de café — sí existe pero no es alimento base
EXCLUIR_ALIMENTOS = {
    # Fauna no disponible
    'ballena (carne)',
    'cobaya',
    'ciervo',
    'corzo',
    # Almidones / ingredientes industriales
    'amidon de arroz',
    'almidon de maiz',
    'almidon de trigo',
    'malta (extracto)',
    'polvo para flanes',
    # Forma concentrada de alimentos básicos (el LP los sobreusa por densidad)
    'huevo (seco)',          # reemplazado por Huevo entero / Clara / Yema crudos
    'leche de vaca (total seca)',   # muy concentrada; leche en polvo descremada sí OK
    'leche de burra',        # no disponible en AR
    # Frutas exóticas / no identificables en góndola
    'majuela',
    # Derivados industriales del aceite
    'aceite higado bacalao',    # suplemento, no aceite de cocina
    # Notas al pie del Excel original — se filtraron como alimentos por error
    'vitamina d',          # líneas de notas sobre vitamina D en unidades internacionales
    # Sustitutos de bebida / no alimentos base
    'achicoria (tostada)',      # sustituto del café, no vegetal de consumo masivo
                                # Además tiene dato de Fe (58 mg/100g) claramente erróneo
    # Legumbres no disponibles en Argentina o peligrosas en exceso
    'almortas',                 # Lathyrus sativus — no disponible en AR; causa latirismo
}

# Exclusión por estado (aplica a CUALQUIER alimento con ese estado)
EXCLUIR_ESTADOS = {
    'almidon',    # patata (almidón) → almidón puro, no papa común
    'diastasada',
}

# ─── Límites individuales por tipo de alimento ────────────────────────────────
# Algunos alimentos son válidos nutricionalmente pero no se consumen en grandes
# cantidades (condimentos, hierbas). El optimizador los sobreusa si no se limitan.
# Formato: {palabra_clave_en_nombre_lower: gramos_maximo_dia}
LIMITES_INDIVIDUALES = {
    # ── Hierbas frescas (condimento, no alimento principal) ───────────────────
    'perejil':     3,    # 1-2 cucharadas máximo
    'cilantro':    3,
    'albahaca':    3,
    'menta':       3,
    'estragon':    3,
    'cebollino':   5,
    'eneldo':      3,
    'laurel':      1,    # hoja de laurel = <1g
    # ── Especias y condimentos secos ─────────────────────────────────────────
    'tomillo':     2,
    'oregano':     2,
    'comino':      2,
    'canela':      3,
    'pimienta':    2,
    'pimenton':    3,
    'azafran':     1,
    'nuez moscada': 1,
    'curry':       3,
    # ── Condimentos líquidos y pastas ─────────────────────────────────────────
    'mostaza':    10,    # 2 cucharaditas
    'ketchup':    30,    # 2 cucharadas
    'mahonesa':   15,    # 1 cucharada
    'mayonesa':   15,
    'vinagre':    15,    # 1 cucharada
    'salsa de soja': 10,
    'salsa worcestershire': 5,
    # ── Otros ingredientes de uso limitado ────────────────────────────────────
    'ajo':        10,    # 2-3 dientes
    'levadura':    7,    # levadura de cerveza o nutricional
    'cacao':      20,    # 2 cucharadas
    'cafe':       15,    # 1-2 tazas
    'te':          5,    # bolsitas
    'miel':       20,    # 1 cucharada
    'alcaparra':   5,
    'aceituna':   30,    # 5-6 aceitunas
    # ── Dulces y golosinas — límite estricto ─────────────────────────────────
    # El LP los elige como fuente barata de HC — limitamos fuertemente
    'caramelo':   10,    # 2-3 caramelos
    'chocolate':  20,    # 2 cuadraditos
    'confitura':  10,    # 1 cucharadita
    'jalea':      10,
    'jaleas':     10,
    'galletita':  30,    # 3-4 galletitas
    'alfajor':    30,
    'facturas':   30,
    'helado':     80,    # 1 bocha pequeña
    'mermelada':  15,    # 1 cucharada
    'dulce de leche': 15,
    'azucar de uva':  10,
    'malta':      10,
    'melaza':     10,
    # ── Otros procesados con azúcar o grasa ──────────────────────────────────
    'galletita':  40,
    'galletas':   60,    # máx ~6 galletas tipo crackers
    'condensada': 30,    # leche condensada: máx 2 cucharadas
    'lacteado':   20,    # bebidas lacteadas azucaradas
    'cereales az': 40,   # cereales azucarados de desayuno
}

# ─── FIX 3: Penalización de formas concentradas/secas ─────────────────────────
# Para alimentos que tienen versión fresca Y seca, el LP tiende a usar la seca
# porque tiene más nutrientes por gramo al mismo precio.
# Solución: multiplicar el precio de la versión seca por un factor de "reconstitución"
# que refleja que 1g de leche seca ≠ 1g de leche fresca para el consumidor.
# El factor es aproximadamente la relación de calorías seco/fresco.
FACTOR_PRECIO_ESTADO = {
    'seco':          3.5,   # deshidratados → caro en equivalente fresco
    'secos':         3.5,
    'seca':          3.5,
    'deshidratado':  3.5,
    'polvo':         4.0,   # leches en polvo, etc.
    'concentrada':   2.5,
    'concent.':      2.5,
    'condens':       2.5,   # condensada
}

# Estados que SÍ son la forma normal de venta (no penalizar)
ESTADOS_NORMALES_SECOS = {
    'frutos secos',  # almendras, nueces, etc. → su forma natural es seca
}

# Grupos donde el estado seco ES la forma habitual de compra
GRUPOS_SECOS_NORMALES = {'FrutosSecos'}



# ─── Diccionario español de España → español rioplatense ─────────────────────
# Se aplica SOLO al campo NOMBRE_COMPLETO (display) — no al campo ALIMENTO
# para no afectar el matching del LP ni el enriquecimiento USDA.
# Orden: más específico primero. Reemplazos por palabra completa (case-insensitive).

DICT_ES_AR = {
    # Vegetales / Hortalizas
    'patata':           'papa',
    'patatas':          'papas',
    'boniato':          'batata',
    'calabacín':        'zapallito',
    'calabacin':        'zapallito',
    'pimiento':         'morrón',
    'pimientos':        'morrones',
    'col de bruselas':  'repollitos de Bruselas',
    'col debruselas':   'repollitos de Bruselas',
    'col rizada':       'kale',
    'col':              'repollo',
    'judías verdes':    'chauchas',
    'judias verdes':    'chauchas',
    'judías blancas':   'porotos blancos',
    'judias blancas':   'porotos blancos',
    'judías rojas':     'porotos rojos',
    'judias rojas':     'porotos rojos',
    'judías':           'porotos',
    'judias':           'porotos',
    'guisantes':        'arvejas',
    'habichuelas':      'chauchas',
    'berberecho':       'berberecho',
    'acelga':           'acelga',
    'escarola':         'escarola',
    'endibia':          'endibia',
    'remolacha':        'remolacha',
    'nabo':             'nabo',
    'chirivía':         'chirivía',
    'chirivia':         'chirivía',
    # Frutas
    'melocotón':        'durazno',
    'melocoton':        'durazno',
    'melocotones':      'duraznos',
    'albaricoque':      'damasco',
    'albaricoques':     'damascos',
    'fresa':            'frutilla',
    'fresas':           'frutillas',
    'fresón':           'frutilla grande',
    'freson':           'frutilla grande',
    'frambuesa':        'frambuesa',
    'ciruela':          'ciruela',
    'aguacate':         'palta',
    'aguacates':        'paltas',
    'cacahuete':        'maní',
    'cacahuetes':       'maní',
    'piñón':            'piñón',
    'níspero':          'níspero',
    'nispero':          'níspero',
    'membrillo':        'membrillo',
    'higo':             'higo',
    'granada':          'granada',
    'pomelo':           'pomelo',
    # Proteínas animales
    'ternera':          'ternera/novillo',
    'bacalao':          'abadejo',
    'rape':             'merluza',
    'lubina':           'róbalo',
    'dorada':           'dorado',
    'gamba':            'langostino',
    'gambas':           'langostinos',
    'langosta':         'langosta',
    'mejillón':         'mejillón',
    'mejillones':       'mejillones',
    'almeja':           'almeja',
    'almejas':          'almejas',
    'calamar':          'calamar',
    'pulpo':            'pulpo',
    'jamón serrano':    'jamón crudo',
    'salchichón':       'salame',
    'butifarra':        'chorizo criollo',
    'morcilla':         'morcilla',
    # Lácteos y derivados
    'nata':             'crema',
    'mantequilla':      'manteca',
    'mahonesa':         'mayonesa',
    'mayonesa':         'mayonesa',
    # Cereales y panificados
    'pan de molde':     'pan lactal',
    'pan de debada':    'pan de cebada',
    'magdalena':        'muffin',
    'magdalenas':       'muffins',

    # Correcciones de estados abreviados o mal escritos del Excel original
    'crdos':           'crudos',
    'crua':            'cruda',
    'frsco':           'fresco',
    'jugop':           'jugo',
    'juego':           'jugo',
    # Nombres mal escritos en el Excel
    'macarroneso fideos': 'fideos macarrones',
    'verduras (medio)':   'verduras mixtas',
    # Cortes de carne con paréntesis redundante
    'Caballo (Carne)':    'Carne de caballo',
    'Cabra (carne)':      'Carne de cabra',
    'Cabrito (carne)':    'Carne de cabrito',
    'Carnero (carne)':    'Carne de carnero',
    'Pollo (corazon)':    'Corazón de pollo',
    'Pollo (higado)':     'Hígado de pollo',
    # Estados numéricos (grado de extracción en harinas)
    '(0.85)':             '(extracción 85%)',
    '(0.80)':             '(extracción 80%)',
    '(0.75)':             '(extracción 75%)',
    '(0.70)':             '(extracción 70%)',
    '(1)':                '(integral)',
    # Abreviaciones del Excel original
    'leche vaca condens.': 'leche condensada',
    'leche vaca cond.':    'leche condensada',
    'leche vaca concent.': 'leche concentrada',
    'leche devaca':        'leche de vaca',
    'leche deoveja':       'leche de oveja',
    'condens.':            'condensada',
    'concent.':            'concentrada',
    'desn.':               'descremada',
    'semimagr.':           'semigrasa',
    # Palabras pegadas sin espacio
    'devaca':              'de vaca',
    'deoveja':             'de oveja',
    'demaiz':              'de maíz',
    # Quesos con nombres poco conocidos en Argentina
    'requeson':            'ricota',
    'queso brugos':        'queso fresco',
    'queso gervaia':       'queso crema',
    'queso emmertal':      'queso gruyere',
    # Otros
    'zumo':             'jugo',
    'maíz':             'maíz',
    'maiz':             'maíz',
    'cacau':            'cacao',
}


def traducir_nombre_ar(nombre: str) -> str:
    """
    Traduce nombres de alimentos de español de España a español rioplatense.
    Aplica reemplazos del diccionario DICT_ES_AR, de mayor a menor longitud.
    """
    resultado = nombre
    for es, ar in sorted(DICT_ES_AR.items(), key=lambda x: len(x[0]), reverse=True):
        low = resultado.lower()
        idx = low.find(es.lower())
        if idx != -1:
            resultado = resultado[:idx] + ar + resultado[idx + len(es):]
    if nombre and nombre[0].isupper() and resultado and resultado[0].islower():
        resultado = resultado[0].upper() + resultado[1:]
    return resultado




# ─── Grupos donde se omite el estado "crudo" en el display ───────────────────
# Nadie come batata cruda, arroz crudo, lentejas crudas, etc.
# El LP usa los valores nutricionales de la tabla (estado crudo como referencia),
# pero mostramos solo el nombre del alimento sin el estado obvio.
# Para carnes, pescados, huevos y lácteos SÍ se mantiene el estado (importa).
GRUPOS_OMITIR_ESTADO_CRUDO = {
    'Hortalizas',   # papa cruda → papa
    'Frutas',       # manzana (cruda) → manzana
    'Leguminosas',  # lentejas (crudas) → lentejas
    'FrutosSecos',  # almendras (seco) → almendras (ya es su forma natural)
    'Cereales',     # arroz (crudo) → arroz
}

# Estados a omitir para esos grupos (solo cuando el nombre resultante sigue siendo claro)
ESTADOS_OMITIR = {'crudo', 'cruda', 'crudos', 'crudas', 'fresco', 'fresca', 'frescos', 'frescas'}



# ─── Alimentos a excluir por no disponibles en Argentina ─────────────────────
ALIMENTOS_EXCLUIR = {
    # Carnes exóticas
    'ballena', 'caballo', 'cabra', 'cabrito', 'carnero',
    # Caza
    'ciervo', 'cobaya', 'conejo', 'corzo', 'jabali', 'jabalí', 'liebre',
    # Aves poco comunes
    'capon', 'capón', 'codoriz', 'codorniz', 'faisan', 'faisán',
    'ganso', 'paloma', 'perdiz', 'pintada',
    'gallina joven', 'gallina vieja',
    # Vísceras y cortes especiales
    'corazon', 'lechecillas', 'higado', 'riñon', 'sesos',
    'tripas', 'sangre', 'lengua',
    # Huevos exóticos
    'huevo de pata', 'huevo de pava', 'huevo seco', 'huevo tortilla',
    # Lácteos no disponibles
    'leche de burra', 'leche de cabra', 'leche de mujer', 'leche de oveja',
    'leche deoveja',
    # Quesos específicos españoles
    'queso brugos', 'queso cabrales', 'queso gervaia', 'queso gorgonzola',
    'queso manchego', 'queso villalon', 'requeson miraflores',
    # Pescados/mariscos exóticos
    'centollo', 'chirla', 'lamprea', 'bogavante', 'breca', 'faneca',
    'gallo', 'barbo', 'carpa', 'congrio', 'hipogloso', 'halibut',
    'caviar', 'cangrejo',
    # Cereales/almidones no comunes
    'sagu', 'sagú', 'tapioca', 'polvo para flanes', 'tarta manzana',
    'pan de viena', 'pan diabeticos',
    # Azúcares/dulces
    'azucar de uva', 'caldo en cubitos', 'malta', 'melaza',
    # Aceites especiales
    'aceite higado bacalao', 'aceite comestible',
    # Embutidos específicos
    'butifarra', 'salchicha de vaca',
    # Oveja
    'carne de oveja',
    # Todo el grupo caza/cordero/ternera
    'cordero', 'ternera',
}

def _esta_excluido_nombre(nombre: str) -> bool:
    """Verifica si un alimento está en la lista de exclusión."""
    n = nombre.lower().strip()
    for excluido in ALIMENTOS_EXCLUIR:
        if excluido in n:
            return True
    return False


def _limpiar_valor(val):
    """Convierte un valor de celda a float, manejando anotaciones como '(1) 4,5'."""
    if val is None:
        return np.nan
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r'^\(\d+\)\s*', '', s)
    s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return np.nan


def _es_excluido(alimento: str, estado: str, grupo: str) -> bool:
    """Determina si un alimento debe marcarse como no disponible."""
    alimento_l = alimento.lower().strip()
    estado_l   = estado.lower().strip() if estado else ''

    # Excluir notas al pie (contienen "vitamina d" o "u.i." o paréntesis de referencia)
    if 'vitamina d' in alimento_l or 'u.i.' in alimento_l:
        return True
    if len(alimento_l) > 80:  # líneas muy largas son notas, no alimentos
        return True

    # Exclusión por nombre completo (nombre + estado)
    nombre_completo = f"{alimento_l} ({estado_l})" if estado_l else alimento_l
    if nombre_completo in EXCLUIR_ALIMENTOS:
        return True
    if alimento_l in EXCLUIR_ALIMENTOS:
        return True

    # Exclusión por estado
    for estado_excl in EXCLUIR_ESTADOS:
        if estado_excl in estado_l:
            return True

    return False


def _factor_precio(estado: str, grupo: str) -> float:
    """Devuelve el multiplicador de precio para estados concentrados."""
    if grupo in GRUPOS_SECOS_NORMALES:
        return 1.0   # frutos secos → no penalizar
    estado_l = estado.lower().strip() if estado else ''
    for estado_key, factor in FACTOR_PRECIO_ESTADO.items():
        if estado_key in estado_l:
            return factor
    return 1.0


def _reclasificar(alimento: str, estado: str, grupo_actual: str) -> str:
    """Aplica correcciones de grupo a alimentos mal clasificados en el Excel."""
    alimento_l = alimento.lower().strip()
    estado_l   = estado.lower().strip() if estado else ''

    for (alim_key, est_key), grupo_correcto in RECLASIFICACIONES.items():
        if alim_key in alimento_l:
            if est_key is None or est_key in estado_l:
                return grupo_correcto
    return grupo_actual


def cargar_tabla(ruta_excel: str | Path) -> pd.DataFrame:
    """
    Carga todas las hojas del Excel y devuelve un DataFrame unificado.

    Returns
    -------
    pd.DataFrame con columnas:
        ALIMENTO, ESTADO, GRUPO, HOJA,
        todas las de NUTRIENT_COLS,
        columnas *_g (por gramo),
        NOMBRE_COMPLETO, DISPONIBLE, FACTOR_PRECIO,
        PRECIO_100G, PRECIO_g (inicialmente NaN)
    """
    ruta = Path(ruta_excel)
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)

    dfs = []

    for hoja, grupo_base in HOJAS_GRUPOS.items():
        if hoja not in wb.sheetnames:
            print(f"⚠ Hoja '{hoja}' no encontrada, saltando.")
            continue

        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))

        if len(filas) < 5:
            continue

        datos = filas[4:]   # primeras 4 filas son encabezados
        registros = []

        for fila in datos:
            if len(fila) < 2 or fila[1] is None:
                continue
            alimento = str(fila[1]).strip()
            if not alimento:
                continue
            # Saltar filas que son sub-títulos de sección
            if alimento.upper() in {
                'ALIMENTOS', 'CEREALES', 'LEGUMINOSAS', 'FRUTOS', 'FRESCOS',
                'SECOS', 'LECHE Y ', 'DERIVADOS', 'HUEVOS', 'AZUCARES',
                'DULCES VARIOS', 'ACEITES Y GRASAS', 'PESCADOS', 'CARNE',
                'CERDO', 'CORDERO', 'OVEJA', 'TERNERA', 'VACA', 'EMBUTIDOS',
                'AVES', 'CAZA', 'HORTALIZAS', 'TUBERCULOS',
            }:
                continue

            estado = str(fila[2]).strip() if fila[2] else ''
            fila_ext = list(fila) + [None] * (33 - len(fila))

            # FIX 1: reclasificación
            grupo = _reclasificar(alimento, estado, grupo_base)

            rec = {
                'NRO':      fila_ext[0],
                'ALIMENTO': alimento,
                'ESTADO':   estado,
                'GRUPO':    grupo,
                'HOJA':     hoja,
            }
            for i, col in enumerate(COL_NAMES[3:], start=3):
                rec[col] = _limpiar_valor(fila_ext[i])

            registros.append(rec)

        if registros:
            dfs.append(pd.DataFrame(registros))

    if not dfs:
        raise ValueError("No se pudo cargar ninguna hoja del Excel.")

    df = pd.concat(dfs, ignore_index=True)

    # ── Conversiones de unidades ──────────────────────────────────────────────
    df['VIT_B1'] = df['VIT_B1'] / 1000    # mcg → mg
    df['VIT_B2'] = df['VIT_B2'] / 1000    # mcg → mg
    df['VIT_C']  = df['VIT_C']  / 1000    # mcg → mg

    # ── FIX 2: columna DISPONIBLE ─────────────────────────────────────────────
    # Aplicar exclusión por nombre
    df['_EXCLUIR_NOMBRE'] = df['ALIMENTO'].apply(_esta_excluido_nombre)

    df['DISPONIBLE'] = df.apply(
        lambda r: not r['_EXCLUIR_NOMBRE'] and not _es_excluido(r['ALIMENTO'], r['ESTADO'], r['GRUPO']),
        axis=1
    )

    # ── Exclusión de ingredientes crudos no comestibles directamente ──────────
    def _es_materia_prima(alimento: str, estado: str) -> bool:
        alimento_l = alimento.lower().strip()
        estado_l   = estado.lower().strip() if estado else ''
        # Por estado
        for est_excl in EXCLUIR_ESTADOS_CRUDOS:
            if est_excl in estado_l:
                return True
        # Por nombre completo
        nombre_completo = f"{alimento_l} ({estado_l})" if estado_l else alimento_l
        if nombre_completo in EXCLUIR_MATERIAS_PRIMAS:
            return True
        if alimento_l in EXCLUIR_MATERIAS_PRIMAS:
            return True
        return False

    mask_mp = df.apply(
        lambda r: _es_materia_prima(r['ALIMENTO'], r['ESTADO']), axis=1
    )
    df.loc[mask_mp, 'DISPONIBLE'] = False

    # ── FIX 3: factor de precio para formas concentradas ─────────────────────
    df['FACTOR_PRECIO'] = df.apply(
        lambda r: _factor_precio(r['ESTADO'], r['GRUPO']),
        axis=1
    )

    # ── Columnas por gramo (para el LP) ──────────────────────────────────────
    for col in NUTRIENT_COLS:
        df[f'{col}_g'] = df[col] / 100.0

    # ── FIBRA estimada (g/100g) ───────────────────────────────────────────────
    def _estimar_fibra(alimento: str, grupo: str) -> float:
        alimento_l = alimento.lower().strip()
        # Buscar override específico por nombre
        for keyword, fibra in FIBRA_ESTIMADA_ALIMENTO.items():
            if keyword in alimento_l:
                return fibra
        # Fallback al promedio del grupo
        return FIBRA_ESTIMADA_GRUPO.get(grupo, 0.0)

    df['FIBRA']   = df.apply(lambda r: _estimar_fibra(r['ALIMENTO'], r['GRUPO']), axis=1)
    df['FIBRA_g'] = df['FIBRA'] / 100.0

    # ── OLIGOELEMENTOS estimados ──────────────────────────────────────────────
    def _estimar_oe(alimento, grupo, por_alimento, por_grupo):
        al = alimento.lower().strip()
        for kw, val in por_alimento.items():
            if kw in al:
                return val
        return por_grupo.get(grupo, 0.0)

    df['ZINC']      = df.apply(lambda r: _estimar_oe(r['ALIMENTO'], r['GRUPO'],
                               ZINC_ESTIMADO_ALIMENTO, ZINC_ESTIMADO_GRUPO), axis=1)
    df['ZINC_g']    = df['ZINC'] / 100.0

    df['YODO']      = df.apply(lambda r: _estimar_oe(r['ALIMENTO'], r['GRUPO'],
                               YODO_ESTIMADO_ALIMENTO, YODO_ESTIMADO_GRUPO), axis=1)
    df['YODO_g']    = df['YODO'] / 100.0

    df['SELENIO']   = df.apply(lambda r: _estimar_oe(r['ALIMENTO'], r['GRUPO'],
                               SELENIO_ESTIMADO_ALIMENTO, SELENIO_ESTIMADO_GRUPO), axis=1)
    df['SELENIO_g'] = df['SELENIO'] / 100.0

    # ── Nombre completo del alimento ──────────────────────────────────────────
    def _nombre_display(row):
        alimento = row['ALIMENTO']
        estado   = str(row['ESTADO']).strip().lower() if row['ESTADO'] else ''
        grupo    = row['GRUPO']

        # Omitir estado obvio (crudo/fresco) para verduras, frutas, cereales, legumbres
        if grupo in GRUPOS_OMITIR_ESTADO_CRUDO and estado in ESTADOS_OMITIR:
            nombre_base = alimento
        elif estado:
            nombre_base = f"{alimento} ({row['ESTADO']})"
        else:
            nombre_base = alimento

        return traducir_nombre_ar(nombre_base)

    df['NOMBRE_COMPLETO'] = df.apply(_nombre_display, axis=1)

    # ── Precio (se completa desde sepa_client.aplicar_precios) ───────────────
    df['PRECIO_100G'] = np.nan
    df['PRECIO_g']    = np.nan

    n_excluidos = (~df['DISPONIBLE']).sum()
    n_total = len(df)
    print(f"✓ Tabla cargada: {n_total} alimentos en {df['GRUPO'].nunique()} grupos")
    print(f"  ({n_total - n_excluidos} disponibles para optimización, "
          f"{n_excluidos} excluidos por baja disponibilidad)")

    # Aplicar enriquecimiento USDA si existe el CSV
    df = aplicar_enriquecimiento_usda(df)

    return df


def resumen_tabla(df: pd.DataFrame) -> None:
    """Imprime un resumen de la tabla cargada."""
    print("\n── Alimentos por grupo (disponibles / total) ────────────────────")
    for grupo, sub in df.groupby('GRUPO'):
        total = len(sub)
        disp  = sub['DISPONIBLE'].sum()
        print(f"  {grupo:<20} {disp:>4} / {total:>4}")
    print(f"  {'TOTAL':<20} {df['DISPONIBLE'].sum():>4} / {len(df):>4}")

    print("\n── Completitud de nutrientes (% filas disponibles con dato) ────")
    df_disp = df[df['DISPONIBLE']]
    for col in NUTRIENT_COLS:
        pct = df_disp[col].notna().mean() * 100
        bar = '█' * int(pct / 5)
        print(f"  {col:<8} {pct:5.1f}% {bar}")

    print(f"\n── Alimentos excluidos ──────────────────────────────────────────")
    excl = df[~df['DISPONIBLE']][['ALIMENTO', 'ESTADO', 'GRUPO']].sort_values('GRUPO')
    print(excl.to_string(index=False))

    print(f"\n── Alimentos con penalización de precio (forma concentrada) ─────")
    pen = df[df['FACTOR_PRECIO'] > 1.0][['NOMBRE_COMPLETO', 'GRUPO', 'FACTOR_PRECIO']]
    print(pen.to_string(index=False))




def aplicar_enriquecimiento_usda(df: pd.DataFrame,
                                  ruta_csv: str | Path = None) -> pd.DataFrame:
    """
    Aplica datos reales de USDA sobre los valores estimados de Zinc y Selenio.
    Si el archivo no existe o falla, mantiene los estimados por grupo.
    """
    if ruta_csv is None:
        ruta_csv = Path(__file__).parent / 'enriquecimiento_usda.csv'

    ruta_csv = Path(ruta_csv)
    if not ruta_csv.exists():
        return df

    try:
        # Leer el CSV detectando separador automáticamente
        # Primero detectar el separador leyendo la primera línea
        with open(ruta_csv, 'r', encoding='utf-8-sig') as f:
            primera_linea = f.readline()
        sep = ';' if primera_linea.count(';') > primera_linea.count(',') else ','

        # Leer como texto puro para evitar cualquier problema de dtype
        df_usda = pd.read_csv(
            ruta_csv,
            sep=sep,
            dtype=str,
            on_bad_lines='skip',
            encoding='utf-8-sig',
        )

        # Limpiar nombres de columnas
        df_usda.columns = [c.strip().upper() for c in df_usda.columns]

        # Convertir columnas numéricas de forma segura
        for col in ['SCORE', 'ZINC_MG', 'YODO_UG', 'SELENIO_UG']:
            if col in df_usda.columns:
                df_usda[col] = pd.to_numeric(
                    df_usda[col].str.replace(',', '.', regex=False),
                    errors='coerce'
                )

        # Filtrar por score mínimo
        if 'SCORE' in df_usda.columns:
            df_usda = df_usda[df_usda['SCORE'] >= 65].copy()

        # Limpiar nombre de alimento
        if 'ALIMENTO' not in df_usda.columns:
            print("⚠ Columna ALIMENTO no encontrada en el CSV")
            return df

        # Construir lookups
        lookup_zinc    = {}
        lookup_selenio = {}

        for _, row in df_usda.iterrows():
            alim = str(row.get('ALIMENTO', '')).strip()
            if not alim:
                continue
            zinc = row.get('ZINC_MG')
            sel  = row.get('SELENIO_UG')
            if pd.notna(zinc):
                try:
                    lookup_zinc[alim] = float(zinc)
                except (ValueError, TypeError):
                    pass
            if pd.notna(sel):
                try:
                    lookup_selenio[alim] = float(sel)
                except (ValueError, TypeError):
                    pass

        # Aplicar al DataFrame usando operaciones vectorizadas
        df = df.copy()

        # Crear Series de lookup mapeando ALIMENTO → valor USDA
        alimentos = df['ALIMENTO'].str.strip()

        zinc_usda    = alimentos.map(lookup_zinc)
        selenio_usda = alimentos.map(lookup_selenio)

        # Reemplazar solo donde hay datos USDA — mantener estimados donde no
        mask_zinc    = zinc_usda.notna()
        mask_selenio = selenio_usda.notna()

        # Forzar columnas a float64 antes de asignar
        df['ZINC']      = df['ZINC'].astype('float64')
        df['ZINC_g']    = df['ZINC_g'].astype('float64')
        df['SELENIO']   = df['SELENIO'].astype('float64')
        df['SELENIO_g'] = df['SELENIO_g'].astype('float64')

        df.loc[mask_zinc,    'ZINC']      = zinc_usda[mask_zinc].astype('float64')
        df.loc[mask_zinc,    'ZINC_g']    = zinc_usda[mask_zinc].astype('float64') / 100.0
        df.loc[mask_selenio, 'SELENIO']   = selenio_usda[mask_selenio].astype('float64')
        df.loc[mask_selenio, 'SELENIO_g'] = selenio_usda[mask_selenio].astype('float64') / 100.0

        print(f"✓ Enriquecimiento USDA aplicado:")
        print(f"  Zinc:    {mask_zinc.sum()} alimentos con datos reales")
        print(f"  Selenio: {mask_selenio.sum()} alimentos con datos reales")
        return df

    except Exception as e:
        print(f"⚠ No se pudo aplicar enriquecimiento USDA: {e}")
        return df

if __name__ == '__main__':
    df = cargar_tabla('tabla_composicion_alimentos.xlsx')
    resumen_tabla(df)

# ─── Exclusión de ingredientes que requieren procesamiento ────────────────────
# Estos alimentos existen en la tabla pero no se consumen directamente —
# son materias primas que necesitan cocción o elaboración industrial.
# El LP los sobreusa porque son baratos y calóricamente densos.
EXCLUIR_ESTADOS_CRUDOS = {
    'grano',      # trigo grano, cebada grano, centeno grano
}

# Alimentos específicos que son materias primas no consumibles directamente
EXCLUIR_MATERIAS_PRIMAS = {
    'harina de trigo',
    'harina de maiz',
    'harina de centeno',
    'harina de cebada',
    'harina de avena',
    'salvado de trigo',
    'salvado de avena',
    'germen de trigo',
    'trigo (grano)',
    'cebada (grano)',
    'centeno (grano)',
    'maiz (grano)',
    'avena (grano)',
}

# ─── Fibra estimada por grupo (g/100g) ───────────────────────────────────────
# La tabla original no tiene fibra. Usamos estimaciones por grupo basadas
# en valores promedio de la base USDA FoodData Central.
# Es una aproximación conservadora — mejor que nada para el LP.
FIBRA_ESTIMADA_GRUPO = {
    'Cereales':    3.5,   # varía mucho: arroz blanco 0.4g, avena 10g → promedio
    'Leguminosas': 8.0,   # lentejas, porotos: 6-10g
    'Hortalizas':  2.5,   # promedio verduras frescas
    'Frutas':      2.0,   # promedio frutas frescas
    'FrutosSecos': 6.5,   # almendras, nueces: 5-8g
    'Lacteos':     0.0,
    'Huevos':      0.0,
    'Aceites':     0.0,
    'Azucares':    0.2,
    'Pescados':    0.0,
    'Carnes':      0.0,
    'Embutidos':   0.0,
    'Aves':        0.0,
}

# Ajuste por alimento específico (override del grupo)
FIBRA_ESTIMADA_ALIMENTO = {
    'arroz':        0.4,
    'arroz blanco': 0.4,
    'papa':         2.2,
    'patata':       2.2,
    'zapallo':      1.0,
    'calabaza':     1.0,
    'banana':       2.6,
    'manzana':      2.4,
    'naranja':      2.4,
    'zanahoria':    2.8,
    'espinaca':     2.2,
    'acelga':       1.6,
    'lechuga':      1.3,
    'tomate':       1.2,
    'lenteja':      8.0,
    'poroto':       7.5,
    'garbanzo':     7.6,
    'soja':         6.0,
    'pan':          2.7,
    'pan integral': 6.0,
    'avena':        10.0,
    'chocolate':    3.4,
    'almendra':     12.5,
    'mani':         8.5,
    'nuez':         6.7,
}

# ─── Oligoelementos estimados (g/100g) ───────────────────────────────────────
# Fuente: USDA FoodData Central SR Legacy + OPS/OMS Tablas de composición
# Unidades: mg/100g para Zinc y Manganeso, µg/100g para Yodo y Selenio
# Estos valores son promedios por grupo — serán reemplazados por valores
# reales cuando se implemente el enriquecimiento desde USDA API (Opción B)

ZINC_ESTIMADO_GRUPO = {       # mg/100g
    'Cereales':    1.0,
    'Leguminosas': 1.5,
    'Hortalizas':  0.3,
    'Frutas':      0.1,
    'FrutosSecos': 3.0,
    'Lacteos':     0.4,
    'Huevos':      1.1,
    'Aceites':     0.0,
    'Azucares':    0.1,
    'Pescados':    0.9,
    'Carnes':      4.5,
    'Embutidos':   2.0,
    'Aves':        2.0,
}

ZINC_ESTIMADO_ALIMENTO = {    # mg/100g — overrides de grupo
    'arroz':       0.6,
    'pan':         0.7,
    'avena':       3.6,
    'lenteja':     3.3,
    'poroto':      3.0,
    'garbanzo':    3.4,
    'almendra':    3.1,
    'mani':        3.3,
    'nuez':        2.7,
    'queso':       3.5,
    'leche':       0.4,
    'pollo':       1.8,
    'bife':        6.3,
    'hígado':      6.0,
    'atun':        0.7,
    'merluza':     0.4,
    'espinaca':    0.5,
    'papa':        0.3,
    'zanahoria':   0.2,
    'tomate':      0.2,
    'banana':      0.2,
    'naranja':     0.1,
}

YODO_ESTIMADO_GRUPO = {       # µg/100g
    'Cereales':    10,
    'Leguminosas':  5,
    'Hortalizas':   3,
    'Frutas':       2,
    'FrutosSecos':  2,
    'Lacteos':     50,    # leche es la principal fuente dietaria de yodo
    'Huevos':      25,
    'Aceites':      1,
    'Azucares':     2,
    'Pescados':   100,    # pescados y mariscos — mayor fuente natural
    'Carnes':       5,
    'Embutidos':    8,
    'Aves':         5,
}

YODO_ESTIMADO_ALIMENTO = {    # µg/100g
    'leche':       44,
    'yogur':       38,
    'queso':       36,
    'merluza':     90,
    'atun':       120,
    'sardina':    135,
    'bacalao':    110,
    'huevo':       26,
    'pan':         12,    # sal yodada usada en panificación
}

SELENIO_ESTIMADO_GRUPO = {    # µg/100g
    'Cereales':   10,
    'Leguminosas': 6,
    'Hortalizas':  1,
    'Frutas':      1,
    'FrutosSecos': 4,
    'Lacteos':     4,
    'Huevos':     20,
    'Aceites':     0,
    'Azucares':    1,
    'Pescados':   30,
    'Carnes':     14,
    'Embutidos':  12,
    'Aves':       17,
}

SELENIO_ESTIMADO_ALIMENTO = { # µg/100g
    'atun':       90,
    'merluza':    36,
    'sardina':    53,
    'pollo':      17,
    'bife':       14,
    'huevo':      20,
    'arroz':       9,
    'avena':      34,
    'nuez brasil':1917,   # la nuez de Brasil es la fuente más rica conocida
}
